"""Export validation_sample.csv to a formatted Excel workbook for Likert annotation."""

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).parent
IN_CSV  = HERE / "outputs" / "validation_sample.csv"
OUT_XLS = HERE / "outputs" / "validation_annotator_1.xlsx"

# Stratum display colours (light fills)
STRATUM_COLOURS = {
    "A_qcet_strong_agree":     "D9EAD3",  # light green
    "B_qcet_strong_rescued":   "B6D7A8",  # medium green
    "C_qcet_partial":          "FFF2CC",  # light yellow
    "D_qcet_disagreement":     "FCE5CD",  # light orange
    "E_aux_specific":          "CFE2F3",  # light blue
    "F_aux_other":             "EAD1DC",  # light pink
    "G_stage3_decisions":      "D9D2E9",  # light purple
    "H_new_qcet_nodes":        "D0E0E3",  # light teal
}

LIKERT_NOTE = (
    "1=Wrong  2=Poor (same area, wrong leaf)  "
    "3=Acceptable (debatable)  4=Good  5=Perfect"
)

# Columns to include and their widths
COLUMNS = [
    ("stratum",          18),
    ("raw_string",       28),
    ("occurrences_llm",  10),
    ("occurrences_human",12),
    ("predicted_id",     14),
    ("predicted_name",   26),
    ("predicted_type",   12),
    ("predicted_source", 20),
    ("predicted_fit",    11),
    ("stage1_qcet_id",   13),
    ("stage1_qcet_fit",  13),
    ("construct",        28),
    ("justification",    55),
    # annotation columns
    ("score",            10),
    ("notes",            40),
]

HEADER_FILL   = PatternFill("solid", fgColor="37474F")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
ANNOT_FILL    = PatternFill("solid", fgColor="FFF9C4")   # pale yellow — annotator fills these
BODY_FONT     = Font(size=10)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
NOWRAP_ALIGN  = Alignment(wrap_text=False, vertical="top")

thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def main() -> None:
    rows = list(csv.DictReader(IN_CSV.open()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation"

    col_keys  = [c[0] for c in COLUMNS]
    col_widths = [c[1] for c in COLUMNS]

    # ── Header row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for ci, (key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=key)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Score column index (1-based)
    score_col_idx = col_keys.index("score") + 1
    score_col_letter = get_column_letter(score_col_idx)

    # ── Data validation: score must be 1–5 ────────────────────────────────────
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        showErrorMessage=True,
        errorTitle="Invalid score",
        error="Enter a value between 1 and 5.",
        showInputMessage=True,
        promptTitle="Likert score",
        prompt=LIKERT_NOTE,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{score_col_letter}2:{score_col_letter}{len(rows)+1}"

    # ── Data rows ──────────────────────────────────────────────────────────────
    prev_stratum = None
    for ri, row in enumerate(rows, start=2):
        stratum = row.get("stratum", "")
        fill_colour = STRATUM_COLOURS.get(stratum, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_colour)

        # slight row-height boost for wrapped text
        ws.row_dimensions[ri].height = 45

        for ci, key in enumerate(col_keys, start=1):
            if key == "score":
                value = None
            elif key == "notes":
                value = ""
            else:
                value = row.get(key, "")
                # convert numeric strings
                if key in ("occurrences_llm", "occurrences_human"):
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass

            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font   = BODY_FONT
            cell.border = THIN_BORDER

            if key == "score":
                cell.fill      = ANNOT_FILL
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "notes":
                cell.fill      = ANNOT_FILL
                cell.alignment = WRAP_ALIGN
            elif key == "justification":
                cell.fill      = row_fill
                cell.alignment = WRAP_ALIGN
            elif key == "raw_string":
                cell.fill      = row_fill
                cell.font      = Font(size=10, bold=True)
                cell.alignment = NOWRAP_ALIGN
            else:
                cell.fill      = row_fill
                cell.alignment = NOWRAP_ALIGN

        # thin separator between strata
        if prev_stratum and stratum != prev_stratum:
            for ci in range(1, len(COLUMNS) + 1):
                c = ws.cell(row=ri, column=ci)
                c.border = Border(
                    left=c.border.left, right=c.border.right,
                    bottom=c.border.bottom,
                    top=Side(style="medium", color="555555"),
                )
        prev_stratum = stratum

    # ── Freeze header + left columns ─────────────────────────────────────────
    ws.freeze_panes = "C2"   # freeze stratum+raw_string columns and header row

    # ── Legend sheet ─────────────────────────────────────────────────────────
    ls = wb.create_sheet("Legend")
    legend_rows = [
        ["Score", "Meaning"],
        [1, "Wrong — incorrect node, different construct"],
        [2, "Poor — same broad area but wrong leaf"],
        [3, "Acceptable — debatable, defensible but not ideal"],
        [4, "Good — correct node, minor quibble"],
        [5, "Perfect — exact match"],
        ["", ""],
        ["Stratum", "Description"],
    ]
    for s, colour in STRATUM_COLOURS.items():
        legend_rows.append([s, ""])
    ls.column_dimensions["A"].width = 30
    ls.column_dimensions["B"].width = 55
    for lr, lrow in enumerate(legend_rows, start=1):
        for lc, val in enumerate(lrow, start=1):
            cell = ls.cell(row=lr, column=lc, value=val)
            cell.font = Font(size=10, bold=(lr in (1, 8)))
    # colour stratum rows in legend
    for i, (s, colour) in enumerate(STRATUM_COLOURS.items()):
        lr = 9 + i
        ls.cell(row=lr, column=1).fill = PatternFill("solid", fgColor=colour)
        ls.cell(row=lr, column=1).value = s

    OUT_XLS.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLS)
    print(f"Wrote {OUT_XLS}  ({len(rows)} rows)")


if __name__ == "__main__":
    main()
