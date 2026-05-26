"""
score_validation.py

Score a Likert-annotated Validation sample.

Reads `outputs/validation_annotator_1.xlsx` after the annotator has
filled the `score` column (1–5) and optional `notes` column.

Score scale:
  5 = Perfect — exact match
  4 = Good    — correct node, minor quibble
  3 = Acceptable — debatable, defensible but not ideal
  2 = Poor    — same broad area but wrong leaf
  1 = Wrong   — incorrect node, different construct

Outputs:
  outputs/validation_summary.md
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from statistics import mean, stdev

import openpyxl

HERE = Path(__file__).parent
DEFAULT_XLS = HERE / "outputs" / "validation_annotator_1.xlsx"
DEFAULT_OUT_MD = HERE / "outputs" / "validation_summary.md"

STRATA_ORDER = [
    "A_qcet_strong_agree",
    "B_qcet_strong_rescued",
    "C_qcet_partial",
    "D_qcet_disagreement",
    "E_aux_specific",
    "F_aux_other",
    "G_stage3_decisions",
    "H_new_qcet_nodes",
]


def fmt_pct(num: int, denom: int) -> str:
    if denom == 0:
        return "n/a"
    return f"{num / denom:.1%}  ({num}/{denom})"


def score_label(s: int) -> str:
    return {1: "Wrong", 2: "Poor", 3: "Acceptable", 4: "Good", 5: "Perfect"}.get(s, "?")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xls", type=Path, default=DEFAULT_XLS)
    ap.add_argument("--out-md", type=Path, default=DEFAULT_OUT_MD)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xls, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        raw_score = r[col["score"]] if "score" in col else None
        if raw_score is None:
            continue
        try:
            score = int(raw_score)
        except (ValueError, TypeError):
            continue
        if score not in (1, 2, 3, 4, 5):
            continue
        rows.append({
            "stratum":          str(r[col["stratum"]] or ""),
            "raw_string":       str(r[col["raw_string"]] or ""),
            "predicted_id":     str(r[col["predicted_id"]] or ""),
            "predicted_name":   str(r[col["predicted_name"]] or ""),
            "predicted_source": str(r[col["predicted_source"]] or ""),
            "score":            score,
            "notes":            str(r[col["notes"]] or "") if "notes" in col else "",
        })

    n_total = ws.max_row - 1
    n_scored = len(rows)
    n_unannotated = n_total - n_scored

    md: list[str] = []
    md.append("# Validation — Likert scorecard")
    md.append("")
    md.append(f"- Sample size:            {n_total}")
    md.append(f"- Annotated rows scored:  {n_scored}")
    md.append(f"- Unannotated (skipped):  {n_unannotated}")
    md.append("")

    if not rows:
        md.append("> No rows have been annotated yet. Fill in the `score` column and re-run.")
        args.out_md.write_text("\n".join(md) + "\n")
        print(f"wrote {args.out_md} (no annotations to score yet)")
        return

    scores = [r["score"] for r in rows]
    n = len(scores)

    n_perfect     = sum(1 for s in scores if s == 5)
    n_good        = sum(1 for s in scores if s == 4)
    n_acceptable  = sum(1 for s in scores if s == 3)
    n_poor        = sum(1 for s in scores if s == 2)
    n_wrong       = sum(1 for s in scores if s == 1)
    n_gte4        = n_perfect + n_good
    n_gte3        = n_gte4 + n_acceptable
    n_lte2        = n_poor + n_wrong

    md.append("## Overall")
    md.append("")
    md.append(f"Mean score: **{mean(scores):.2f}** / 5.0"
              + (f"  (SD {stdev(scores):.2f})" if n > 1 else ""))
    md.append("")
    md.append("| score | label | count | % |")
    md.append("|---|---|---|---|")
    for val, label in [(5, "Perfect"), (4, "Good"), (3, "Acceptable"), (2, "Poor"), (1, "Wrong")]:
        cnt = sum(1 for s in scores if s == val)
        md.append(f"| {val} | {label} | {cnt} | {cnt/n:.1%} |")
    md.append("")
    md.append("| threshold | value |")
    md.append("|---|---|")
    md.append(f"| Score ≥ 4 (Good or Perfect)       | {fmt_pct(n_gte4, n)} |")
    md.append(f"| Score ≥ 3 (Acceptable or better)  | {fmt_pct(n_gte3, n)} |")
    md.append(f"| Score ≤ 2 (Poor or Wrong)         | {fmt_pct(n_lte2, n)} |")
    md.append(f"| Score = 1 (Wrong)                 | {fmt_pct(n_wrong, n)} |")
    md.append("")

    md.append("## Per-stratum breakdown")
    md.append("")
    md.append("| stratum | n | mean | ≥4 | ≥3 | ≤2 |")
    md.append("|---|---|---|---|---|---|")
    by_stratum: dict[str, list[int]] = defaultdict(list)
    for r in rows:
        by_stratum[r["stratum"]].append(r["score"])
    for stratum in STRATA_ORDER:
        ss = by_stratum.get(stratum, [])
        if not ss:
            continue
        ns = len(ss)
        md.append(
            f"| `{stratum}` | {ns} | {mean(ss):.2f} "
            f"| {fmt_pct(sum(1 for s in ss if s >= 4), ns)} "
            f"| {fmt_pct(sum(1 for s in ss if s >= 3), ns)} "
            f"| {fmt_pct(sum(1 for s in ss if s <= 2), ns)} |"
        )
    md.append("")

    md.append("## Per-source breakdown")
    md.append("")
    md.append("| chosen_source | n | mean | ≥4 | ≥3 | ≤2 |")
    md.append("|---|---|---|---|---|---|")
    by_source: dict[str, list[int]] = defaultdict(list)
    for r in rows:
        by_source[r["predicted_source"]].append(r["score"])
    for source in sorted(by_source):
        ss = by_source[source]
        ns = len(ss)
        md.append(
            f"| `{source}` | {ns} | {mean(ss):.2f} "
            f"| {fmt_pct(sum(1 for s in ss if s >= 4), ns)} "
            f"| {fmt_pct(sum(1 for s in ss if s >= 3), ns)} "
            f"| {fmt_pct(sum(1 for s in ss if s <= 2), ns)} |"
        )
    md.append("")

    low_rows = [r for r in rows if r["score"] <= 2]
    if low_rows:
        md.append(f"## Low-scoring rows (score ≤ 2, {len(low_rows)} rows)")
        md.append("")
        md.append("| score | stratum | raw_string | predicted_id | predicted_name | notes |")
        md.append("|---|---|---|---|---|---|")
        for r in sorted(low_rows, key=lambda x: x["score"]):
            md.append(
                f"| {r['score']} | `{r['stratum']}` | `{r['raw_string']}` "
                f"| {r['predicted_id']} | {r['predicted_name'][:30]} "
                f"| {(r['notes'] or '').replace('|', '/')[:80]} |"
            )
        md.append("")

    args.out_md.write_text("\n".join(md) + "\n")
    print(f"wrote {args.out_md}")
    print(f"overall mean score: {mean(scores):.2f} / 5.0  (n={n})")
    print(f"score ≥ 4: {fmt_pct(n_gte4, n)}   score ≤ 2: {fmt_pct(n_lte2, n)}")


if __name__ == "__main__":
    main()
