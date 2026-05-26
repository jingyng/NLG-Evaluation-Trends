#!/usr/bin/env python3
"""Convert polysemous_mappings_review.csv into a formatted Excel workbook
for manual annotation.

Sheet 1 (Review): the 127 mappings, with frozen header, autofilter, sized
columns, and conditional highlighting on the bare polysemous strings (the
single-word forms most likely to need a different target).

Sheet 2 (Targets): the 119 candidate target ids (117 QCET leaves + the two
auxiliary categories) with names + short definitions, so reviewers can
pick `proposed_qcet_id` without having to look elsewhere.

Run from `paper_code/`:

    python 05_criteria_normalization/build_polysemous_review_xlsx.py
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
CSV_PATH = HERE / "polysemous_mappings_review.csv"
QCET_JSON = HERE / "qcet_taxonomy.json"
XLSX_PATH = HERE / "polysemous_mappings_review.xlsx"

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
ANNOTATION_FILL = PatternFill("solid", fgColor="FFF2CC")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FCE4D6")
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))


DIVERGED_FILL = PatternFill("solid", fgColor="F4B084")          # orange — Pass-1 vs stage-4 diverged


def write_review_sheet(ws, rows: list[dict]) -> None:
    headers = [
        "root_word", "raw_lowercase", "raw_string",
        "occurrences_llm", "occurrences_human", "total_occ",
        "pass1_qcet_id", "pass1_qcet_name", "pass1_fit",
        "stage4_qcet_id", "stage4_qcet_name", "stage4_source",
        "diverged", "proposed_qcet_id", "notes",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["root_word"], r["raw_lowercase"], r["raw_string"],
            int(r["occurrences_llm"]), int(r["occurrences_human"]), int(r["total_occ"]),
            r["pass1_qcet_id"], r["pass1_qcet_name"], r["pass1_fit"],
            r["stage4_qcet_id"], r["stage4_qcet_name"], r["stage4_source"],
            r["diverged"], r["proposed_qcet_id"], r["notes"],
        ])

    # Header style
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER

    # Body style
    annotation_cols = {"proposed_qcet_id", "notes"}
    annotation_idx = {headers.index(c) + 1 for c in annotation_cols}
    diverged_idx = headers.index("diverged") + 1
    for row_idx in range(2, ws.max_row + 1):
        # Bare polysemous strings: lowercase form == root word.
        bare = ws.cell(row=row_idx, column=1).value == ws.cell(row=row_idx, column=2).value
        # Pass-1 disagreed with the force-merge winner.
        diverged = bool(ws.cell(row=row_idx, column=diverged_idx).value)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in annotation_idx:
                c.fill = ANNOTATION_FILL
            elif diverged:
                c.fill = DIVERGED_FILL
            elif bare:
                c.fill = HIGHLIGHT_FILL

    # Column widths
    widths = {
        "root_word": 14, "raw_lowercase": 28, "raw_string": 30,
        "occurrences_llm": 12, "occurrences_human": 14, "total_occ": 10,
        "pass1_qcet_id": 11, "pass1_qcet_name": 32, "pass1_fit": 9,
        "stage4_qcet_id": 12, "stage4_qcet_name": 32, "stage4_source": 18,
        "diverged": 9, "proposed_qcet_id": 18, "notes": 36,
    }
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[h]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_targets_sheet(ws, qcet_taxonomy: dict) -> None:
    headers = ["qcet_id", "name", "level1", "level2", "level3", "short_definition"]
    ws.append(headers)

    leaves = [n for n in qcet_taxonomy["nodes"] if n["is_leaf"]]
    leaves.sort(key=lambda n: (n.get("level1", ""), n.get("level2", "") or "", n["id"]))
    for leaf in leaves:
        ws.append([
            leaf["id"], leaf["name"],
            leaf.get("level1", ""), leaf.get("level2", "") or "",
            leaf.get("level3", "") or "",
            (leaf.get("short_definition") or "").strip().replace("\n", " "),
        ])

    # Two aux entries
    ws.append(["AUX-OverallQuality", "Overall Quality / Preference",
               "(meta)", "(meta)", "(meta)",
               "Holistic preference judgements that resist QCET decomposition."])
    ws.append(["AUX-Other", "Other / Unclassifiable",
               "(meta)", "(meta)", "(meta)",
               "Free-form fragments / metric names / out-of-scope strings; dropped from analyses."])

    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {"qcet_id": 12, "name": 44, "level1": 8, "level2": 8,
              "level3": 8, "short_definition": 72}
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[h]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_instructions_sheet(ws) -> None:
    instructions = [
        "How to use this workbook",
        "",
        "1. Each row in the 'Review' sheet is ONE case-sensitive raw variant of a polysemous criterion,",
        "   showing its independent Pass-1 verdict alongside the eventual stage-4 force-merged target.",
        "   Case variants of the same lowercase form appear adjacent (sorted by lowercase, then occ).",
        "",
        "2. Cell shading:",
        "   - ORANGE row: Pass-1 disagreed with the force-merge winner (the case-collapse step at",
        "     stage 4 overwrote this variant's Pass-1 target with the dominant variant's target).",
        "     These are the most likely to deserve a per-variant override.",
        "   - PINK row: bare polysemous string (lowercase form == root word, no qualifier).",
        "     Often where the wrong eventual target lives even when no Pass-1 divergence flagged it.",
        "   - YELLOW cells: your annotation columns (proposed_qcet_id, notes).",
        "",
        "3. To override a mapping: type the new target id (e.g., QOC-w-1, QIC-w-1, QEC-c-2)",
        "   in the 'proposed_qcet_id' column. Leave blank to keep the current stage-4 mapping.",
        "   The 'Targets' sheet has all 119 valid ids with definitions.",
        "",
        "4. Use 'notes' to record reasoning (e.g., \"bare form; for MT/QA papers should be task-specific\").",
        "",
        "Pre-flagged candidates (highest paper-impact first):",
        "",
        "  - 'accuracy' (~361 occ): bare form currently QEC-c-1 Factual Truth. Likely should be",
        "    QOC-w-1 Correctness of Outputs (generic correctness, no qualifier).",
        "  - 'consistency' (~254 occ): bare form QOG-c-4 Internal Consistency of Outputs. Could be",
        "    QIC-c-3 Consistency with Input depending on intended sense.",
        "  - 'faithfulness' (~212 occ): bare form QIC-c-3 Consistency with Input. WALKTHROUGH default",
        "    for summarization is QEC-c-2 Relative Factual Accuracy.",
        "  - 8 rows are flagged ORANGE: those are the diverged Pass-1 vs stage-4 cases.",
        "",
        "After annotation, hand the file back; we'll apply the proposed_qcet_id values via",
        "apply_polysemous_overrides.py and re-run the downstream pipeline.",
    ]
    for line in instructions:
        ws.append([line])
    ws.column_dimensions["A"].width = 110
    for row_idx in range(1, ws.max_row + 1):
        c = ws.cell(row=row_idx, column=1)
        if row_idx == 1 or "priority" in str(c.value).lower():
            c.font = Font(bold=True, size=13)
        c.alignment = Alignment(wrap_text=True, vertical="center")


def main() -> None:
    if not CSV_PATH.exists():
        raise SystemExit(f"CSV not found: {CSV_PATH} — run list_polysemous_mappings.py first.")

    with CSV_PATH.open(newline="") as f:
        rows = list(csv.DictReader(f))

    qcet_taxonomy = json.load(QCET_JSON.open())

    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    write_instructions_sheet(ws_inst)

    ws_review = wb.create_sheet("Review")
    write_review_sheet(ws_review, rows)

    ws_targets = wb.create_sheet("Targets")
    write_targets_sheet(ws_targets, qcet_taxonomy)

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")
    print(f"  Sheet 'Review':       {len(rows)} mappings to annotate")
    print(f"  Sheet 'Targets':      reference for the 119 valid target ids")
    print(f"  Sheet 'Instructions': how-to + priority-order suggestion")


if __name__ == "__main__":
    main()
