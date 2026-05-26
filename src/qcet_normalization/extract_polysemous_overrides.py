#!/usr/bin/env python3
"""Read the annotated polysemous_mappings_review.xlsx and write the overrides
out as a CSV that apply_qcet_to_metadata.py can consume.

The overrides apply at the lowercase-key level (so all case variants of a
raw string are remapped together), matching how normalize_merged_results.py
performs its mapping lookup.

Run from `paper_code/`:

    python 05_criteria_normalization/extract_polysemous_overrides.py
"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "polysemous_mappings_review.xlsx"
QCET_JSON = HERE / "qcet_taxonomy.json"
OUT_CSV = HERE / "polysemous_overrides.csv"


# AUX targets that aren't in qcet_taxonomy.json's leaves but are valid targets.
_AUX_NAMES = {
    "AUX-OverallQuality": "Overall Quality / Preference",
    "AUX-Other":          "Other / Unclassifiable",
}


def _load_id_to_name() -> dict[str, str]:
    data = json.load(QCET_JSON.open())
    out = {n["id"]: n["name"] for n in data["nodes"] if n["is_leaf"]}
    out.update(_AUX_NAMES)
    return out


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Annotated workbook not found: {XLSX}")

    id_to_name = _load_id_to_name()

    wb = load_workbook(XLSX, data_only=True)
    if "Review" not in wb.sheetnames:
        raise SystemExit("Sheet 'Review' missing from workbook")
    ws = wb["Review"]
    headers = [c.value for c in ws[1]]
    required = {"raw_lowercase", "raw_string", "stage4_qcet_id", "proposed_qcet_id", "notes"}
    missing = required - set(headers)
    if missing:
        raise SystemExit(f"Missing columns in Review sheet: {missing}")

    by_lower: dict[str, dict] = {}              # raw_lowercase → override entry
    conflicts: list[tuple[str, set[str]]] = []
    invalid: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        prop = (rec.get("proposed_qcet_id") or "")
        if isinstance(prop, str):
            prop = prop.strip()
        if not prop:
            continue

        raw_lower = (rec.get("raw_lowercase") or "").strip().lower()
        if not raw_lower:
            continue

        if prop not in id_to_name:
            invalid.append((rec.get("raw_string") or "", prop))
            continue

        existing = by_lower.get(raw_lower)
        if existing and existing["new_qcet_id"] != prop:
            conflicts.append((raw_lower, {existing["new_qcet_id"], prop}))
            continue

        by_lower[raw_lower] = {
            "raw_lowercase": raw_lower,
            "old_qcet_id":   rec.get("stage4_qcet_id") or "",
            "new_qcet_id":   prop,
            "new_qcet_name": id_to_name[prop],
            "notes":         (rec.get("notes") or "").strip(),
            # Carry one example case-form for traceability.
            "example_case_form": rec.get("raw_string") or "",
        }

    if invalid:
        print("ERROR: invalid proposed_qcet_id values (not in QCET 117 leaves or AUX):")
        for raw, prop in invalid:
            print(f"  {raw!r} → {prop!r}")
        sys.exit(2)

    if conflicts:
        print("ERROR: conflicting overrides for the same lowercase key:")
        for raw, ids in conflicts:
            print(f"  {raw!r} has competing overrides: {sorted(ids)}")
        sys.exit(2)

    fieldnames = [
        "raw_lowercase", "old_qcet_id", "new_qcet_id", "new_qcet_name",
        "example_case_form", "notes",
    ]
    rows = sorted(by_lower.values(), key=lambda d: d["raw_lowercase"])
    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    print(f"Wrote {len(rows)} overrides to {OUT_CSV}")
    for r in rows:
        print(f"  {r['raw_lowercase']:35s}  {r['old_qcet_id']:20s} → {r['new_qcet_id']}  ({r['new_qcet_name']})")


if __name__ == "__main__":
    main()
